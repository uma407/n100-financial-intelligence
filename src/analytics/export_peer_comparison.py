import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font


DB_PATH = "nifty100.db"
OUTPUT_PATH = Path("output/peer_comparison.xlsx")


def load_data():
    conn = sqlite3.connect(DB_PATH)

    percentiles = pd.read_sql_query(
        "SELECT * FROM peer_percentiles",
        conn,
    )

    ratios = pd.read_sql_query(
        "SELECT * FROM financial_ratios",
        conn,
    )

    companies = pd.read_sql_query(
        "SELECT id AS company_id, company_name FROM companies",
        conn,
    )

    conn.close()

    return percentiles, ratios, companies


def build_peer_sheet(peer_group, percentiles, ratios, companies):
    peer_data = percentiles[percentiles["peer_group_name"] == peer_group]

    if peer_data.empty:
        return pd.DataFrame()

    pivot = peer_data.pivot_table(
        index=["company_id", "year"],
        columns="metric",
        values="percentile_rank",
        aggfunc="mean",
    ).reset_index()

    metric_values = ratios.merge(
        companies,
        on="company_id",
        how="left",
    )

    result = metric_values.merge(
        pivot,
        on=["company_id", "year"],
        how="inner",
        suffixes=("", "_percentile"),
    )

    return result


def color_percentile_cells(workbook):
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gold = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]

        percentile_cols = [
            idx + 1
            for idx, header in enumerate(headers)
            if header in [
                "roe",
                "net_profit_margin_pct",
                "debt_to_equity",
                "revenue_cagr_5yr",
                "pat_cagr_5yr",
                "eps_cagr_5yr",
                "interest_coverage",
                "asset_turnover",
                "composite_quality_score",
            ]
        ]

        for row in sheet.iter_rows(min_row=2):
            for col_idx in percentile_cols:
                cell = row[col_idx - 1]
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 75:
                        cell.fill = green
                    elif cell.value <= 25:
                        cell.fill = red
                    else:
                        cell.fill = yellow

        # Highlight first data row as benchmark row
        if sheet.max_row >= 2:
            for cell in sheet[2]:
                cell.fill = gold
                cell.font = Font(bold=True)


def add_summary_rows(writer, sheet_name, df):
    workbook = writer.book
    sheet = workbook[sheet_name]

    summary_row = sheet.max_row + 2
    sheet.cell(row=summary_row, column=1, value="Peer Group Median")

    for col_idx, col in enumerate(df.columns, start=1):
        if pd.api.types.is_numeric_dtype(df[col]):
            median_value = df[col].median()
            sheet.cell(row=summary_row, column=col_idx, value=median_value)


def export_peer_comparison():
    OUTPUT_PATH.parent.mkdir(exist_ok=True)

    percentiles, ratios, companies = load_data()

    peer_groups = [
        group
        for group in percentiles["peer_group_name"].dropna().unique()
        if group != "No peer group assigned"
    ]

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        for peer_group in sorted(peer_groups):
            df = build_peer_sheet(
                peer_group,
                percentiles,
                ratios,
                companies,
            )

            if df.empty:
                continue

            if "composite_quality_score" in df.columns:
                df = df.sort_values(
                    "composite_quality_score",
                    ascending=False,
                )

            sheet_name = peer_group[:31]

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

            add_summary_rows(writer, sheet_name, df)

        color_percentile_cells(writer.book)

    print(f"Peer comparison report saved: {OUTPUT_PATH}")
    print("Sheets generated:", len(peer_groups))


if __name__ == "__main__":
    export_peer_comparison()