from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

from engine import preset_screen


OUTPUT_PATH = Path("output/screener_output.xlsx")

PRESETS = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
    "turnaround_watch",
]


def clean_sheet_name(name):
    return name[:31]


def export_screener_output():
    OUTPUT_PATH.parent.mkdir(exist_ok=True)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        for preset in PRESETS:
            df = preset_screen(preset)

            if "composite_quality_score" in df.columns:
                df = df.sort_values("composite_quality_score", ascending=False)

            required_cols = [
                "company_id",
                "year",
                "roe",
                "debt_to_equity",
                "net_profit_margin_pct",
                "operating_profit_margin_pct",
                "asset_turnover",
                "free_cash_flow_cr",
                "capex_cr",
                "earnings_per_share",
                "cash_from_operations_cr",
                "revenue_cagr_5yr",
                "pat_cagr_5yr",
                "eps_cagr_5yr",
                "fcf_conversion_rate_pct",
                "composite_quality_score",
            ]

            available_cols = [c for c in required_cols if c in df.columns]
            df_export = df[available_cols].head(50)

            sheet_name = clean_sheet_name(preset)
            df_export.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = green_fill
                        elif cell.value < 0:
                            cell.fill = red_fill

    print(f"Screener output saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    export_screener_output()